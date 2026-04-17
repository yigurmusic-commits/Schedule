"""
Экспорт расписания в Excel.
Упрощенная версия под новую схему.
"""

from io import BytesIO
from typing import List, Dict, Tuple, Optional
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.models.models import (
    ScheduleEntry, Group, TimeSlot, DayOfWeek,
    Subject, Teacher, Classroom
)

# Стили
thin_side = Side(style="thin")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_TITLE = Font(bold=True, color="FFFFFF")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def export_schedule_to_excel(db: Session) -> BytesIO:
    # Загружаем всё текущее расписание
    entries = db.query(ScheduleEntry).all()
    groups = db.query(Group).order_by(Group.name).all()
    time_slots = db.query(TimeSlot).order_by(TimeSlot.slot_number).all()
    days = db.query(DayOfWeek).order_by(DayOfWeek.id).all()
    
    subjects_map = {s.id: s.name for s in db.query(Subject).all()}
    teachers_map = {t.id: t.full_name for t in db.query(Teacher).all()}
    classrooms_map = {c.id: c.room_number for c in db.query(Classroom).all()}

    # Карта: group_id -> {(day_id, slot_id): entry}
    sched: Dict[int, Dict[Tuple[int, int], ScheduleEntry]] = {g.id: {} for g in groups}
    for e in entries:
        sched[e.group_id][(e.day_id, e.slot_id)] = e

    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    # Заголовки
    ws.cell(row=1, column=1, value="День").font = FONT_TITLE
    ws.cell(row=1, column=1).fill = FILL_HEADER
    ws.cell(row=1, column=2, value="Пара").font = FONT_TITLE
    ws.cell(row=1, column=2).fill = FILL_HEADER
    
    col = 3
    group_col_map = {}
    for g in groups:
        cell = ws.cell(row=1, column=col, value=g.name)
        cell.font = FONT_TITLE
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        group_col_map[g.id] = col
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 25
        col += 1

    current_row = 2
    for d in days:
        day_start_row = current_row
        for ts in time_slots:
            ws.cell(row=current_row, column=1, value=d.name).alignment = ALIGN_CENTER
            ws.cell(row=current_row, column=2, value=str(ts.slot_number)).alignment = ALIGN_CENTER
            
            for g in groups:
                entry = sched[g.id].get((d.id, ts.id))
                if entry:
                    subj = subjects_map.get(entry.subject_id, "—")
                    teacher = teachers_map.get(entry.teacher_id, "—")
                    room = classrooms_map.get(entry.classroom_id, "—")
                    text = f"{subj}\n{teacher}\nкаб. {room}"
                    cell = ws.cell(row=current_row, column=group_col_map[g.id], value=text)
                    cell.alignment = ALIGN_CENTER
                    cell.font = Font(size=8)
                
                ws.cell(row=current_row, column=group_col_map[g.id]).border = thin_border
            
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=2).border = thin_border
            current_row += 1
        
        # Объединяем ячейки дня
        if current_row - 1 > day_start_row:
            ws.merge_cells(start_row=day_start_row, start_column=1, end_row=current_row - 1, end_column=1)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
